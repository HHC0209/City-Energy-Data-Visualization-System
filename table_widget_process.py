from PyQt5 import QtCore, QtGui, QtWidgets
import xlwt
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
# from xlwt import Workbook


class TableWidgetProcess:
    def __init__(self, widget):
        self.datas = []
        self.selected_year = []
        self.widget = widget
        self.mode = 0  # 0 forecast 1 analyze
        self.path_name = []


    def clear(self):
        self.widget.clear()
        self.datas = []
        self.selected_year = []
        self.widget.setRowCount(0)
        self.widget.setColumnCount(0)


    def add_data(self, datas, selected_years):
        self.datas = datas
        self.selected_year = selected_years
        self.set_path_name()


    def set_path_name(self):
        self.path_name = []
        for data in self.datas:
            name = '数据详情：'
            name += data['级别1'] + ' > ' + data['级别2'] + ' > ' + data['级别3'] + ' > ' + data['级别4'] + ' > ' + data['级别5'] + ' > ' + data['级别6']
            self.path_name.append(name)


    def display_forecast(self, results, pmax):
        self.mode = 0  # mode等于0 意味着是预测模式
        year = [int(item) for item in self.selected_year]
        year.sort()
        start_year = year[-1] + 1
        self.header = list(self.datas[0].keys())[1:] + ['%d(预测)' % year for year in range(start_year, pmax+1)]
        self.widget.setRowCount(len(self.datas))
        self.widget.setColumnCount(len(self.header))
        self.widget.setHorizontalHeaderLabels(self.header)
        self.widget.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.ResizeToContents)
        row = 0
        for point in self.datas:
            col = 0
            for key in self.header[:7]:
                item = QtWidgets.QTableWidgetItem(point[key])
                item.setTextAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter)
                self.widget.setItem(row, col, item)
                col += 1
            for key in self.header[7:7 + len(self.selected_year)]:
                if point[key] != None:
                    item = QtWidgets.QTableWidgetItem(str(round(float(point[key]), 2)))
                    item.setTextAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter)
                    self.widget.setItem(row, col, item)
                col += 1
            for col_forecast in range(len(results[0])):
                if results[row][col_forecast] != None:
                    item = QtWidgets.QTableWidgetItem(str(round(results[row][col_forecast], 2)))
                    item.setTextAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter)
                    # print(row, col + col_forecast, results[row][col_forecast], item)
                    self.widget.setItem(row, col+col_forecast, item)
            row += 1


    def display_analyze(self, results, headers = None):
        self.set_path_name()
        self.mode = 1 # mode等于1意味着是分析模式
        self.widget.setRowCount(len(self.datas))
        self.widget.setColumnCount(len(self.datas))
        if headers == None:
            self.header = [(str(data['编码'])+'\n'+data['地域']) for data in self.datas]
            self.widget.setHorizontalHeaderLabels(self.header)
            self.widget.setVerticalHeaderLabels(self.header)

        else:
            self.header = None
            self.header_h = headers[0]
            self.header_v = headers[1]
            self.widget.setColumnCount(len(headers[0]))
            self.widget.setRowCount(len(headers[1]))
            self.widget.setHorizontalHeaderLabels(self.header_h)
            self.widget.setVerticalHeaderLabels(self.header_v)

        for i in range(len(self.header)):
            self.widget.horizontalHeaderItem(i).setToolTip(self.path_name[i])
            self.widget.verticalHeaderItem(i).setToolTip(self.path_name[i])

        self.widget.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.ResizeToContents)
        self.widget.verticalHeader().setSectionResizeMode(QtWidgets.QHeaderView.ResizeToContents)
        self.result = results

        cmap = plt.cm.get_cmap('coolwarm')
        row = 0
        for result_row in results:
            col = 0
            for result in result_row:
                try:
                    result = float(result)
                except:
                    result = 0
                result_norm = round(result/2 + 0.5, 3)
                item = QtWidgets.QTableWidgetItem(str(round(result, 3)))
                item.setTextAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter)
                if self.header != None:
                    if str(result) != 'nan' and row != col:
                        try:
                            r, g, b, a = cmap(result_norm)
                            item.setBackground(QtGui.QColor(int(r*255), int(g*255), int(b*255)))
                        except:
                            pass
                    
                self.widget.setItem(row, col, item)

                col += 1
            row += 1


    def export_excel(self, filepath):
        # 存在问题：模式0使用xls进行保存
        # 模式1使用xlsx进行保存
        # 解决方法：返回table_widget的模式
        workbook = xlwt.Workbook(encoding='utf-8')
        worksheet = workbook.add_sheet('Data')

        style_hv = xlwt.XFStyle()
        font = xlwt.Font()
        font.name = '微软雅黑'

        alignment_hv = xlwt.Alignment()
        alignment_hv.horz = xlwt.Alignment.HORZ_CENTER
        alignment_hv.vert = xlwt.Alignment.VERT_CENTER

        style_hv.alignment = alignment_hv
        style_hv.font = font

        if self.mode == 0:
            for i, h in enumerate(self.header):
                worksheet.write(0, i, h, style_hv)
            for row in range(len(self.datas)):
                for col in range(len(self.header)):
                    if self.widget.item(row, col) != None:
                        data = self.widget.item(row, col).text()
                    else:
                        data = ""
                    try:
                        data = float(data)
                    except ValueError:
                        pass
                    worksheet.write(row + 1, col, data, style_hv)

            worksheet.col(0).width = 3000
            worksheet.col(1).width = 3000
            for col in range(2, 8):
                worksheet.col(col).width = 6200
            for col in range(8, len(self.datas)):
                worksheet.col(col).width = 3000
            workbook.save(filepath)
        else:
            cmap = plt.cm.get_cmap('coolwarm')
            workbook = Workbook()
            worksheet = workbook.active
            if self.header != None:
                header_horizon = self.header
                header_vertical = self.header
            else:
                header_horizon = self.header_h
                header_vertical = self.header_v

            for col in range(len(header_horizon)):
                cell = worksheet.cell(row=1, column=col+2, value=header_horizon[col])
                cell.font = Font(name='微软雅黑')
                # cell.font = xlwt.Font(name='微软雅黑')
            for row in range(len(header_vertical)):
                cell = worksheet.cell(row=row+2, column=1, value=header_vertical[row])
                cell.font = Font(name='微软雅黑')
                # cell.font = xlwt.Font(name='微软雅黑')
            if self.header != None:
                for x in range(len(header_vertical)):
                    for y in range(len(header_horizon)):
                        result = str(round(self.result[x][y], 3))
                        cell = worksheet.cell(row=x+2, column=y+2, value=result)
                        cell.font = Font(name='微软雅黑')
                        # cell.font = xlwt.Font(name='微软雅黑')
                        if result != 'nan' and x != y:
                            result_norm = round((self.result[x][y]+1)/2, 3)
                            r, g, b, a = cmap(result_norm)
                            r_hex = hex(int(255*r))[2:].rjust(2, '0')
                            g_hex = hex(int(255*g))[2:].rjust(2, '0')
                            b_hex = hex(int(255*b))[2:].rjust(2, '0')
                            color = r_hex + g_hex + b_hex
                            cell.fill = PatternFill(fgColor=color, fill_type="solid")
                            # cell.fill = xlwt.Pattern(fgColor=color, fill_type="solid")
            else:
                for x in range(len(header_vertical)):
                    for y in range(len(header_horizon)):
                        try:
                            result = round(float(self.result[x][y]), 3)
                        except ValueError:
                            result = 0
                        cell = worksheet.cell(row=x+2, column=y+2, value=result)
                        cell.font = Font(name='微软雅黑')
                        # cell.font = xlwt.Font(name='微软雅黑')

            workbook.save(filepath)


    def getMode(self):
        return self.mode